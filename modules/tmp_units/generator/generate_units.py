#!/usr/bin/python
 
###  Copyright 2019 Thomas Figueroa.
###
###  Distributed under the Boost Software License, Version 1.0.
###
###  See accompanying file LICENSE_1_0.txt or copy at
###  http://www.boost.org/LICENSE_1_0.txt

# TODO: Radians / Steradians shouldn't get prefixes
# TODO: Units that do not begin with the unit, but are prefixed already,
#        like square or cubic, need to be identified so that the prefix
#        is appended in the appropriate place.
# TODO: Degrees Celsius should be spelled out have the prefix before
#        the 'degree' part

import os
import pathlib
from openpyxl import load_workbook

SI_Derived_Units_dict = {}
SI_Base_Units_dict = {}
prefixes_dict = {}

def loadUnitWorkbook():
    unit_wb = load_workbook(filename='./Units_Library.xlsx',read_only=True,data_only=True)

    prefix_sheet = unit_wb['SI_Prefixes']
    SI_Derived_Units_sheet = unit_wb['SI_Derived_Units']
    SI_Base_Units = unit_wb['SI_Base_Units']

    # Decrease to min_row=2 and delete max_row when zepto, yocto, zetta, and yotta are available
    # in the std::ratio library
    for row in prefix_sheet.iter_rows(min_row=4, max_row=20, values_only=True):
        prefix_id = row[0]
        prefix = {
            "Symbol": row[1],
            "Factor": row[2],
            "Value": row[3]
        }
        prefixes_dict[prefix_id] = prefix

    for row in SI_Derived_Units_sheet.iter_rows(min_row=2, values_only=True):
        unit_name = row[0]
        unit_ = {
            "Name": row[0],
            "Type": row[1],
            "Symbol": row[2],
            "Pretty": row[3],
            "Impl_Num": row[4],
            "Impl_Den": row[5]
        }
        SI_Derived_Units_dict[unit_name] = unit_

    for row in SI_Base_Units.iter_rows(min_row=2, values_only=True):
        unit_name = row[0]
        unit_ = {
            "Name": row[0],
            "Type": row[1],
            "Symbol": row[2],
            "Pretty": row[3],
            "Impl_Num": row[4],
            "Impl_Den": row[5]
        }
        SI_Base_Units_dict[unit_name] = unit_

def generateHeader(author_name):
    return \
        f"//  Copyright 2019-2022 {author_name}.\n" \
        f"//\n" \
        f"//  Distributed under the Boost Software License, Version 1.0.\n" \
        f"//\n" \
        f"//  See accompanying file LICENSE_1_0.txt or copy an\n" \
        f"//  http://www.boost.org/LICENSE_1_0.txt\n" \
        f"\n"

def generateUnitModuleFiles(unit_name):
    return \
        f"export module Boost.TMP.Units.Unit.{unit_name.title()};\n\n" \
        f"import Boost.TMP;\n" \
        f"import Boost.TMP.Units.Engine.Base;\n" \
        f"\n" \
        f"import std.core;\n" \
        f"\n"

def generateUnitBeginNamespace(unit_type, impl_numer, impl_denom):
    if impl_numer is None:
        impl_numer = ""
    if impl_denom is None:
        impl_denom = ""

    return \
        f"namespace tmp = boost::tmp;\n" \
        f"export namespace potato::units {{\n" \
        f"\n" \
        f"{' ':>4}using {unit_type}_tag_t = tmp::list_<tmp::list_<{impl_numer}>, tmp::list_<{impl_denom}>>;\n" \
        f"\n" \
        f"{' ':>4}template<typename T>\n" \
        f"{' ':>4}concept {unit_type.title()}C = std::same_as<typename T::impl, {unit_type}_tag_t>;\n" \
        f"\n"

def generateUnitStruct(unit_name, pretty, impl_numer, impl_denom):
    if impl_numer is None:
        impl_numer = ""
    if impl_denom is None:
        impl_denom = ""
    if pretty is None:
        pretty = "N/A"

    return \
        f"{' ':>4}template <typename RatioTypeT, typename P>\n" \
        f"{' ':>4}struct {unit_name}_impl {{\n" \
        f"\n" \
        f"{' ':>8}static constexpr auto pretty = \"{pretty}\";\n" \
        f"{' ':>8}using DerivedValueType = tmp::call_<\n" \
        f"{' ':>12}tmp::if_<tmp::is_<std::uint64_t>,\n" \
        f"{' ':>16}tmp::always_<std::int64_t>,\n" \
        f"{' ':>16}tmp::always_<long double>\n" \
        f"{' ':>12}>, P>;\n\n" \
        f"{' ':>8}using mod_ratio  = RatioTypeT;\n" \
        f"{' ':>8}using value_type = DerivedValueType;\n" \
        f"{' ':>8}using numer_type = DerivedValueType;\n" \
        f"{' ':>8}using impl       = {unit_name}_tag_t;\n" \
        f"\n" \
        f"{' ':>8}constexpr {unit_name}_impl(value_type val) : value{{val}} {{}}\n\n" \
        f"{' ':>8}using is_any_impl = tmp::false_;\n" \
        f"\n" \
        f"{' ':>8}value_type value{{}};\n" \
        f"{' ':>4}}};\n" \
        f"\n"

def generatePrefixedUnit(unit_name, unit_prefix, unit_suffix, unit_numerical_type):
    if unit_prefix is None:
        unit_prefix = ""
        unit_prefix_type_ = f"unity_ratio"
    else:
        unit_prefix_type_ = f"std::{unit_prefix}"

    # split the string before per
    unit_split = unit_name.split('per')

    unit_pre = ""
    
    if 'square' in unit_split[0]:
        old_square = "square_"
        new_square = "square_"+unit_prefix
        unit_pre = unit_name.replace(old_square, new_square)
    elif 'cubic' in unit_split[0]:
        old_square = "cubic_"
        new_square = "cubic_"+unit_prefix
        unit_pre = unit_name.replace(old_square, new_square)
    elif 'inverse' in unit_split[0]:
        old_square = "inverse_"
        new_square = "inverse_"+unit_prefix
        unit_pre = unit_name.replace(old_square, new_square)
    else:
        unit_pre = f"{unit_prefix}{unit_name}"
    return \
        f"{' ':>4}using {unit_pre}{unit_suffix} {'=':^8} {unit_name}_impl<{unit_prefix_type_}, {unit_numerical_type}>;\n"

def generateUnitOperator(unit_name, unit_prefix, unit_suffix, unit_numerical_type, unit_prefix_symbol, unit_symbol):
    if unit_prefix is None:
        unit_prefix = ""
    if unit_prefix_symbol is None:
        unit_prefix_symbol = ""

    # split the string before per
    unit_split = unit_name.split('per')

    unit_pre = ""
    
    if 'square' in unit_split[0]:
        old_square = "square_"
        new_square = "square_"+unit_prefix
        unit_name = unit_name.replace(old_square, new_square)
        old_sq_unit = "sq_"
        new_sq_unit = "sq_"+unit_prefix_symbol
        unit_symbol_mod = unit_symbol.replace(old_sq_unit, new_sq_unit)
    elif 'cubic' in unit_split[0]:
        old_square = "cubic_"
        new_square = "cubic_"+unit_prefix
        unit_name = unit_name.replace(old_square, new_square)
        old_cu_unit = "cubic_"
        new_cu_unit = "cubic_"+unit_prefix_symbol
        unit_symbol_mod = unit_symbol.replace(old_cu_unit, new_cu_unit)
    elif 'inverse' in unit_split[0]:
        old_square = "inverse_"
        new_square = "inverse_"+unit_prefix
        unit_name = unit_name.replace(old_square, new_square)
        old_inv_unit = "inverse_"
        new_inv_unit = "inverse_"+unit_prefix_symbol
        unit_symbol_mod = unit_symbol.replace(old_inv_unit, new_inv_unit)
    else:
        unit_name = f"{unit_prefix}{unit_name}"
        unit_symbol_mod = f"{unit_prefix_symbol}{unit_symbol}"
    return \
        f"{' ':>4}constexpr {unit_name}{unit_suffix} operator\"\"_{unit_symbol_mod}({unit_numerical_type} val) {{\n" \
        f"{' ':>8}return static_cast<{unit_name}{unit_suffix}>(val);\n" \
        f"{' ':>4}}}\n"

def generateUnitEndNamespace():
    return \
        f"}} // namespace potato::units\n"

def generateUnitFile(unit_name, unit_type, unit_symbol="", pretty="", impl_numer="", impl_denom=""):
    with open(f"{unit_type}.ixx", "w", encoding="utf8") as file:
        file.write(generateHeader("Thomas Figueroa")) # For now.
        file.write(generateUnitModuleFiles(unit_name))
        file.write(generateUnitBeginNamespace(unit_name, impl_numer, impl_denom))
        file.write(generateUnitStruct(unit_name, pretty, impl_numer, impl_denom))
        for prefix_, vals in prefixes_dict.items():
            file.write(generatePrefixedUnit(unit_name, prefix_, "_ld", "long double"))
        for prefix_, vals in prefixes_dict.items():
            file.write(generatePrefixedUnit(unit_name, prefix_, "_ull", "unsigned long long"))
        file.write('\n')
        for prefix_, vals in prefixes_dict.items():
            file.write(generateUnitOperator(unit_name, prefix_, "_ld", "long double", vals['Symbol'], unit_symbol))
        for prefix_, vals in prefixes_dict.items():
            file.write(generateUnitOperator(unit_name, prefix_, "_ull", "unsigned long long", vals['Symbol'], unit_symbol))
        file.write('\n')
        file.write(generateUnitEndNamespace())

def generateDispatcherIncludeFiles():
    return \
        f"\n" \
        f"export module Boost.TMP.Units.Engine.TypeDispatcher;\n\n" \
        f"import Boost.TMP;\n" \
        f"\n" \
        f"import Boost.TMP.Units;\n" \
        f"import Boost.TMP.Units.Engine.Base;\n" \
        f"import Boost.TMP.Units.Anything;\n" \
        f"\n" \
        f"import std.core;\n" \
        f"\n"

def generateDispatcherBeginNamespace():
    return \
        f"export namespace potato::units {{\n" \
        f"{' ':>4}using namespace boost::tmp;\n" \
        f"\n"

def generateDispatcherMainStruct():
    return \
        f"{' ':>4}template <typename L, typename... Ls>\n" \
        f"{' ':>4}struct dispatcher;\n" \
        f"\n"

# TODO: Multiple types for the same basic type (length: imperial, SI)
def generateDispatcherStruct(unit_name, unit_type, unit_symbol="", pretty="", impl_numer="", impl_denom=""):
    if impl_numer is None:
        impl_numer = ""
    if impl_denom is None:
        impl_denom = ""
    
    return \
        f"{' ':>4}// {unit_type}\n" \
        f"{' ':>4}template <>\n" \
        f"{' ':>4}struct dispatcher<list_<list_<{impl_numer}>, list_<{impl_denom}>>> {{\n" \
        f"{' ':>8}template <typename T, typename P>\n" \
        f"{' ':>8}using f = {unit_name}_impl<T, P>;\n" \
        f"{' ':>4}}};\n" \
        f"\n"

def generateUnitlessStruct():
    return \
        f"{' ':>4}// Unitless\n" \
		f"{' ':>4}template <>\n" \
		f"{' ':>4}struct dispatcher<list_<list_<>, list_<>>> {{\n" \
		f"{' ':>8}template <typename T, typename P>\n" \
		f"{' ':>8}using f = P;\n" \
		f"{' ':>4}}};\n" \
        f"\n"

def generateAnythingStructs():
    return \
        f"{' ':>4}template <typename Anything>\n" \
        f"{' ':>4}struct dispatcher<Anything> {{\n" \
        f"{' ':>8}template <typename T, typename P>\n" \
        f"{' ':>8}using f = anything_impl<T, P, Anything>;\n" \
        f"{' ':>4}}};\n" \
        f"\n" \
        f"{' ':>4}template <typename Anything>\n" \
        f"{' ':>4}struct dispatcher<list_<Anything>> {{\n" \
        f"{' ':>8}template <typename T, typename P>\n" \
        f"{' ':>8}using f = anything_impl<T, P, Anything>;\n" \
        f"{' ':>4}}};\n" \
        f"\n"

def generateTypeDispatcher():
    with open(f"type_dispatcher.ixx", "w", encoding="utf8") as file:
        file.write(generateHeader("Thomas Figueroa"))
        file.write(generateDispatcherIncludeFiles())
        file.write(generateDispatcherBeginNamespace())
        file.write(generateDispatcherMainStruct())
        file.write(generateUnitlessStruct())
        for keys, vals in SI_Base_Units_dict.items():
            file.write(generateDispatcherStruct(*vals.values()))
        for keys, vals in SI_Derived_Units_dict.items():
            file.write(generateDispatcherStruct(*vals.values()))
        file.write(generateAnythingStructs())
        file.write(generateUnitEndNamespace())

def generateUnitModuleImport(unit_name, unit_type, unit_symbol="", pretty="", impl_numer="", impl_denom=""):
    return \
        f"export import Boost.TMP.Units.Unit.{unit_name.title()};\n"

def generateAnythingImplType():
    return \
        f"{' ':>4}template <typename T, typename P, typename Impl>\n" \
        f"{' ':>4}struct anything_impl;\n" \
        f"\n"

def generateAllUnitsModule():
    with open(f"tmp_all_units.ixx", "w", encoding="utf8") as file:
        file.write(generateHeader("Thomas Figueroa"))
        file.write(f"export module Boost.TMP.Units;\n\n")
        for keys, vals in SI_Base_Units_dict.items():
            file.write(generateUnitModuleImport(*vals.values()))
        for keys, vals in SI_Derived_Units_dict.items():
            file.write(generateUnitModuleImport(*vals.values()))

loadUnitWorkbook()

pathlib.Path('../units/SI/base').mkdir(parents=True, exist_ok=True)
pathlib.Path('../units/SI/derived').mkdir(parents=True, exist_ok=True)
pathlib.Path('../engine').mkdir(parents=True, exist_ok=True)

si_base_path = pathlib.Path('../units/SI/base')
si_derived_path = pathlib.Path('../units/SI/derived')
units_engine = pathlib.Path('../engine')
generator_directory = os.getcwd()

# change to SI base directory
os.chdir(si_base_path)
for keys, vals in SI_Base_Units_dict.items():
    generateUnitFile(*vals.values())
os.chdir(generator_directory)

# change to SI derived directory
os.chdir(si_derived_path)
for keys, vals in SI_Derived_Units_dict.items():
    generateUnitFile(*vals.values())
os.chdir(generator_directory)

# change to engine directory
os.chdir(units_engine)
generateTypeDispatcher()
generateAllUnitsModule()
